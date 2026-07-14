import openpyxl
import pandas as pd
import math
import os
import logging
from typing import List, Dict, Any, Tuple, Generator
from datetime import datetime, date

logger = logging.getLogger("app.services.excel")

class ExcelService:
    @staticmethod
    def inspect_workbook(file_path: str) -> Dict[str, Any]:
        """
        Profiles the Excel workbook. Returns metadata and sheets info.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found at {file_path}")
            
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            file_size = os.path.getsize(file_path)
            
            return {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "file_size_bytes": file_size,
                "sheet_names": sheet_names,
                "created_at": datetime.utcnow().isoformat()
            }
        except Exception as e:
            logger.error(f"Failed to inspect workbook: {str(e)}")
            raise ValueError(f"Invalid Excel file: {str(e)}")

    @classmethod
    def profile_sheet(cls, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Detailed profiling of a single sheet:
        Finds header row, column list, inferred types, total rows, and returns sample rows.
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet {sheet_name} not found in workbook")
            
        sheet = wb[sheet_name]
        
        # 1. Detect header row
        header_row_index, original_headers = cls._detect_header_row(sheet)
        if not original_headers:
            wb.close()
            return {
                "sheet_name": sheet_name,
                "is_empty": True,
                "total_rows": 0,
                "columns": []
            }
            
        # Normalize headers
        normalized_headers = [cls._normalize_header(h) for h in original_headers]
        
        # 2. Extract preview rows (up to 100 rows) and count total rows
        preview_rows = []
        total_rows = 0
        
        # Read sheet row by row starting after header_row_index
        row_count = 0
        for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            # Check if row is completely empty
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
                
            total_rows += 1
            if len(preview_rows) < 100:
                # Align row values to headers (truncate or pad if mismatched)
                aligned_row = list(row[:len(original_headers)])
                while len(aligned_row) < len(original_headers):
                    aligned_row.append(None)
                
                # Convert date objects to string for JSON serialization
                serialized_row = []
                for val in aligned_row:
                    if isinstance(val, (datetime, date)):
                        serialized_row.append(val.isoformat())
                    elif isinstance(val, float) and math.isnan(val):
                        serialized_row.append(None)
                    else:
                        serialized_row.append(val)
                        
                preview_rows.append(dict(zip(normalized_headers, serialized_row)))
                
        # 3. Infer column types based on first 500 rows
        sample_rows_for_type = []
        sheet_iter = sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + 501, values_only=True)
        for row in sheet_iter:
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            aligned_row = list(row[:len(original_headers)])
            while len(aligned_row) < len(original_headers):
                aligned_row.append(None)
            sample_rows_for_type.append(aligned_row)
            
        inferred_types = cls._infer_column_types(normalized_headers, sample_rows_for_type)
        
        # Create columns list with metadata
        columns = []
        for orig, norm in zip(original_headers, normalized_headers):
            columns.append({
                "original_name": orig,
                "normalized_name": norm,
                "inferred_type": inferred_types[norm]
            })
            
        wb.close()
        
        return {
            "sheet_name": sheet_name,
            "is_empty": False,
            "header_row_index": header_row_index,
            "total_rows": total_rows,
            "columns": columns,
            "sample_data": preview_rows[:10]  # Just 10 rows for preview inside the columns schema list
        }

    @classmethod
    def read_sheet_chunks(cls, file_path: str, sheet_name: str, header_row_index: int, columns: List[str], chunk_size: int = 1000) -> Generator[List[Dict[str, Any]], None, None]:
        """
        Generator yielding chunks of rows from a sheet. Memory-conscious.
        """
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[sheet_name]
        
        chunk = []
        for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
                
            aligned_row = list(row[:len(columns)])
            while len(aligned_row) < len(columns):
                aligned_row.append(None)
                
            # Create dict matching headers
            row_dict = {}
            for col, val in zip(columns, aligned_row):
                if isinstance(val, float) and math.isnan(val):
                    row_dict[col] = None
                else:
                    row_dict[col] = val
            
            chunk.append(row_dict)
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
                
        if chunk:
            yield chunk
            
        wb.close()

    @staticmethod
    def _detect_header_row(sheet) -> Tuple[int, List[str]]:
        """
        Heuristically detects which row contains the headers.
        Checks the first 10 rows. The row containing the most non-empty unique string values is chosen.
        """
        best_row_idx = 1
        best_headers = []
        max_score = -1
        
        # Check first 10 rows
        for r_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            # Calculate score for this row
            cells = [c for c in row if c is not None]
            if not cells:
                continue
                
            # Filter for string cells
            string_cells = [str(c).strip() for c in cells if str(c).strip() != ""]
            unique_string_cells = set(string_cells)
            
            # Simple score: count of unique non-empty string values
            # Deduct points if values look like numbers
            score = 0
            for val in string_cells:
                # Is it a header candidate? Strings are good, numbers are suspicious
                try:
                    float(val)
                    # Numeric cell, less likely to be a header
                    score += 0.1
                except ValueError:
                    # Non-numeric string, good header candidate
                    score += 1.0
                    
            # Normalize by unique count to penalize rows with lots of duplicates (e.g. repeated data)
            if len(string_cells) > 0:
                score *= (len(unique_string_cells) / len(string_cells))
                
            if score > max_score and len(unique_string_cells) > 0:
                max_score = score
                best_row_idx = r_idx
                best_headers = [str(c).strip() for c in row if c is not None]
                
        return best_row_idx, best_headers

    @staticmethod
    def _normalize_header(header: str) -> str:
        """
        Normalizes Excel column name into standard snake_case representation.
        """
        if not header:
            return "unnamed_column"
        norm = str(header).strip().lower()
        norm = norm.replace(" ", "_").replace("-", "_").replace(".", "_").replace("/", "_")
        norm = "".join([c for c in norm if c.isalnum() or c == "_"])
        # Ensure it starts with a letter or underscore
        if norm and norm[0].isdigit():
            norm = f"col_{norm}"
        return norm if norm else "unnamed_column"

    @staticmethod
    def _infer_column_types(headers: List[str], sample_rows: List[List[Any]]) -> Dict[str, str]:
        """
        Infers types based on checking values in the sample rows.
        """
        inferred = {h: "String" for h in headers}
        if not sample_rows:
            return inferred
            
        for col_idx, norm_header in enumerate(headers):
            col_values = []
            for row in sample_rows:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and str(val).strip() != "":
                        col_values.append(val)
                        
            if not col_values:
                inferred[norm_header] = "String"
                continue
                
            # Tally types
            types_count = {"Int": 0, "Float": 0, "DateTime": 0, "Boolean": 0, "String": 0}
            for val in col_values:
                # Check for datetime/date objects
                if isinstance(val, (datetime, date)):
                    types_count["DateTime"] += 1
                    continue
                # Check Boolean
                if isinstance(val, bool) or str(val).strip().lower() in ("true", "false", "yes", "no"):
                    types_count["Boolean"] += 1
                    continue
                # Check Int
                try:
                    int(str(val))
                    types_count["Int"] += 1
                    continue
                except ValueError:
                    pass
                # Check Float
                try:
                    float(str(val))
                    types_count["Float"] += 1
                    continue
                except ValueError:
                    pass
                
                types_count["String"] += 1
                
            # Determine majority type
            # Find key with max count
            max_type = max(types_count, key=types_count.get)
            
            # Map simplified types to standard ClickHouse compatible types
            if max_type == "Int":
                inferred[norm_header] = "Int64"
            elif max_type == "Float":
                inferred[norm_header] = "Float64"
            elif max_type == "DateTime":
                inferred[norm_header] = "DateTime"
            elif max_type == "Boolean":
                inferred[norm_header] = "UInt8"  # ClickHouse represents Bool as UInt8 (0 or 1) or Boolean
            else:
                inferred[norm_header] = "String"
                
        return inferred
